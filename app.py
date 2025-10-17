import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Function to format the packing list
def format_packing_list(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Packing List"

    # Header Section
    ws.merge_cells("A1:J1")
    ws["A1"] = "PACKING LIST"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:B2")
    ws["A2"] = "Customer Name:"
    ws.merge_cells("C2:F2")

    ws.merge_cells("A3:B3")
    ws["A3"] = "Country:"
    ws.merge_cells("C3:F3")

    ws.merge_cells("A4:B4")
    ws["A4"] = "Customer Cd:"
    ws.merge_cells("C4:F4")
    ws["C4"] = df["Customer"].iloc[0] if "Customer" in df.columns else ""

    ws.merge_cells("G2:G3")
    ws["G2"] = "PACKING LIST NO:"
    ws.merge_cells("H2:J3")
    if "Document No" in df.columns:
        ws["H2"] = ", ".join(df["Document No"].dropna().unique().astype(str))

    ws["G4"] = "Date:"
    ws.merge_cells("H4:J4")
    ws["H4"] = datetime.today().strftime("%d/%m/%Y")

    # Apply borders to header
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=10):
        for cell in row:
            cell.border = thin_border

    # Table Section
    headers = ["Sno", "Part No", "Brand", "Manf.Part", "Qty", "Unit Price (in AED)",
               "Amount (in AED)", "Description", "COO", "HSCODE"]
    ws.append(headers)

    for i, row in df.iterrows():
        ws.append([
            i + 1,
            row.get("Part No", ""),
            row.get("Brand", ""),
            row.get("Manf.Part", ""),
            row.get("Qty Avl", ""),
            row.get("Price", ""),
            row.get("Amount", ""),
            row.get("Description", ""),
            row.get("COO", ""),
            row.get("HSCODE", "")
        ])

    # Summary Section
    start_row = ws.max_row + 2
    summary_labels = ["TOTAL AMOUNT", "DISCOUNT", "NET AMOUNT", "VAT 5%", "TOTAL AED"]
    amount_total = df["Amount"].sum() if "Amount" in df.columns else 0
    discount = 0
    net_amount = amount_total - discount
    vat = net_amount * 0.05
    total_aed = net_amount + vat
    summary_values = [amount_total, discount, net_amount, vat, total_aed]

    for i, label in enumerate(summary_labels):
        row_num = start_row + i
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        ws.cell(row=row_num, column=1).value = label
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_num, column=7).value = summary_values[i]
        for col in range(1, 11):
            ws.cell(row=row_num, column=col).border = thin_border

    # Signature Section
    sig_start = ws.max_row + 2
    ws.merge_cells(start_row=sig_start, start_column=1, end_row=sig_start + 4, end_column=10)
    ws.cell(row=sig_start, column=1).value = "AUTHORIZED SIGNATORY"
    ws.cell(row=sig_start, column=1).alignment = Alignment(horizontal="left", vertical="bottom")
    for row in ws.iter_rows(min_row=sig_start, max_row=sig_start + 4, min_col=1, max_col=10):
        for cell in row:
            cell.border = thin_border

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    return output

# Streamlit UI
st.title("Packing List Formatter")
uploaded_file = st.file_uploader("Upload Dump Excel File", type=["xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file, engine="openpyxl")
    formatted_file = format_packing_list(df)
    st.download_button("Download Formatted Packing List", formatted_file.getvalue(), file_name="Formatted_Packing_List.xlsx")
